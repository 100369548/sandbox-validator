#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通讯沙盘校验工具 v5.0 (Python GUI)
基于 Go v4.8 完整移植所有校验规则
支持批量处理 + 实时进度 + 错误导出
"""

import os
import re
import sys
import time
import json
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter


# ==================== 校验规则定义 ====================

# 通讯通路门店字段定义 (与 Go v4.8 完全一致)
STORE_FIELDS = [
    (0, "省份", "enum", True, {"海南省","广东省","广西壮族自治区","云南省","四川省","河北省","山西省","辽宁省","吉林省","黑龙江省","江苏省","浙江省","安徽省","福建省","江西省","山东省","河南省","湖北省","湖南省","海南省","重庆市","新疆维吾尔自治区","宁夏回族自治区","内蒙古自治区","广西壮族自治区","贵州省","甘肃省","青海省","贵州省","吉林省","黑龙江省","北京市","上海市","天津市"}),
    (1, "城市", "text", True, None),
    (2, "区县", "text", True, None),
    (3, "乡镇", "text", True, None),
    (4, "进度", "enum", True, {"已存在","无规划","洽谈中"}),
    (5, "编码", "code", True, None),
    (6, "名称", "text", True, None),
    (7, "地理编码", "ref", True, "geo"),
    (8, "商圈编码", "ref", True, "biz"),
    (9, "类型", "enum", True, {"综合卖场","荣耀体验店","荣耀授权专卖店","耀生活","华为授权体验店","华为堡垒店","华为非授权店","OPPO体验店","VIVO体验店","小米体验店","苹果体验店","其他品牌专卖店","运营商厅店","NKA综合卖场","其他"}),
    (10, "capa", "uint", True, None),
    (11, "面积", "uint", False, None),
    (12, "状态", "enum", True, {"不合作","营业中","闭店","装修中"}),
    (13, "运营商", "enum", True, {"公开","公开&三网","排它-移动","排它-电信","排它-联通","非排它-移动电信","非排它-移动联通","非排它-电信联通"}),
    (14, "联系人", "text", True, None),
    (15, "联系电话", "phone", True, None),
    (16, "地址", "text", True, None),
    (17, "渠道编码", "chan", True, None),
    (18, "客户名称", "text", True, None),
    (19, "0-2500", "uint", True, None),
    (20, "2500-4000", "uint", True, None),
    (21, ">4000", "uint", True, None),
    (22, "荣耀", "uint", True, None),
    (23, "华为", "uint", True, None),
    (24, "小米", "uint", True, None),
    (25, "VIVO", "uint", True, None),
    (26, "苹果", "uint", True, None),
    (27, "OPPO", "uint", True, None),
    (28, "荣耀PC", "uint", True, None),
    (29, "华为PC", "uint", True, None),
    (30, "荣耀平板", "uint", True, None),
    (31, "华为平板", "uint", True, None),
    (32, "PC", "uint", True, None),
    (33, "平板", "uint", True, None),
    (34, "手表", "uint", True, None),
    (35, "荣耀阵地数", "uint", False, None),
    (36, "荣耀阵地位", "pos", False, None),
    (37, "荣耀门头数", "uint", False, None),
    (38, "华为阵地数", "uint", False, None),
    (39, "华为阵地位", "pos", False, None),
    (40, "华为门头数", "uint", False, None),
    (41, "小米阵地数", "uint", False, None),
    (42, "小米阵地位", "pos", False, None),
    (43, "小米门头数", "uint", False, None),
    (44, "VIVO阵地数", "uint", False, None),
    (45, "VIVO阵地位", "pos", False, None),
    (46, "VIVO门头数", "uint", False, None),
    (47, "苹果阵地数", "uint", False, None),
    (48, "苹果阵地位", "pos", False, None),
    (49, "苹果门头数", "uint", False, None),
    (50, "OPPO阵地数", "uint", False, None),
    (51, "OPPO阵地位", "pos", False, None),
    (52, "OPPO门头数", "uint", False, None),
    (53, "荣耀促销员", "uint", True, None),
    (54, "华为促销员", "uint", True, None),
    (55, "小米促销员", "uint", True, None),
    (56, "VIVO促销员", "uint", True, None),
    (57, "苹果促销员", "uint", True, None),
    (58, "OPPO促销员", "uint", True, None),
    (59, "客户员工", "uint", True, None),
    (60, "店端人数", "uint", True, None),
    (61, "国补", "enum", True, {"是","否"}),
    (62, "城乡类别", "enum", True, {"主城区","非主城区"}),
    (63, "街边/MALL", "enum", True, {"街边店","MALL店"}),
    (64, "备注", "text", False, None),
]

# 通讯客户沙盘字段定义 (只校验 E/F/I-T/AQ)
CUSTOMER_FIELDS = [
    (4, "跨省份", "enum", True, {"是","否"}),
    (5, "跨地市", "enum", True, {"是","否"}),
    (8, "圈层", "text", True, None),
    (9, "华为属性", "enum", True, {"金种子","TOP368","TOP1000","TOP1500","NDKA+","TOP2000","TOP5000","其他"}),
    (10, "OPPO层级", "enum", True, {"T+","T","S","A","其他"}),
    (11, "vivo层级", "enum", True, {"黑金+","黑金","钻石","铂金","其他"}),
    (12, "小米层级", "enum", True, {"蓝血Ultra","蓝血Plus","蓝血","金牌","卓越Ultra","卓越Plus","卓越","优秀","其他"}),
    (13, "运营商属性", "enum", True, {"公开","公开&三网","排它-移动","排它-电信","排它-联通","非排它-移动电信","非排它-移动联通","非排它-电信联通"}),
    (14, "友商代理", "enum", True, {"是","否"}),
    (15, "车授权", "enum", True, {"是","否"}),
    (16, "老板名", "text", True, None),
    (17, "老板电话", "phone", True, None),
    (18, "操盘手名", "text", True, None),
    (19, "操盘手电话", "phone", True, None),
    (42, "0-2500容量", "uint", True, None),
]


# ==================== 引用数据 ====================
class RefData:
    def __init__(self):
        self.geo_codes = set()
        self.biz_codes = set()


# ==================== 校验函数 ====================

def to_int(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s in ('', 'nan', 'N/A', 'None'):
        return None
    s = s.replace(',', '')
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return None

def is_blank(val):
    if val is None:
        return True
    return str(val).strip() == ''

def safe_str(val):
    if val is None:
        return ''
    s = str(val).strip()
    if s.lower() in ('nan', 'none', 'n/a'):
        return ''
    return s

def clean_phone(v):
    return re.sub(r'[^0-9]', '', str(v))


def chk_enum(name, val, eset):
    if is_blank(val):
        return False, "该字段必填"
    sv = safeStr(val)
    if sv in eset:
        return True, None
    return False, f"无效值: {'/'.join(sorted(eset))}"

def chk_uint(name, val, req):
    if is_blank(val):
        if req:
            return False, "该字段必填"
        return True, None
    iv = to_int(val)
    if iv is None:
        return False, "非有效整数"
    if iv < 0 or iv > 999999:
        return False, f"超出范围: {iv}"
    return True, None

def chk_phone(name, val, req):
    if is_blank(val):
        if req:
            return False, "该字段必填"
        return True, None
    sv = safeStr(val)
    if sv == '无':
        return True, None
    s = cleanPhone(sv)
    if re.match(r'^1[3-9]\d{9}$', s):
        return True, None
    return False, "无效手机号"

def chk_code(name, val, plan, req):
    if is_blank(val):
        if req and plan == '已存在':
            return False, '已存在门店必须填写编码'
        return True, None
    s = safeStr(val)
    if s in ('无规划', '洽谈中', '/', '无'):
        return True, None
    if re.match(r'^[A-Za-z]', s) or re.match(r'^46\d{6,}$', s):
        return True, None
    return False, '编码格式异常'

def chk_pos(name, val):
    if is_blank(val):
        return True, None
    sv = safeStr(val)
    if sv in ('0', '0.0', ''):
        return True, None
    if sv in ('TOP1', 'TOP2', 'TOP3', '其他', '无陣地'):
        return True, None
    return False, '无效枚举: TOP1/TOP2/TOP3/其他/无陣地'

def chk_ref(name, val, ref_set, req, label):
    if is_blank(val):
        if req:
            return False, '该字段必填'
        return True, None
    sv = safeStr(val)
    if sv in ref_set:
        return True, None
    return False, f'{label}编码不存在于系统'

def chk_chan(name, val, req):
    if is_blank(val):
        if req:
            return False, '该字段必填'
        return True, None
    s = safeStr(val)
    if s == '无' or re.match(r'^\d+$', s):
        return True, None
    return False, "渠道编码应为纯数字或'无'"


# ==================== 单行校验 ====================

def validate_store_row(row, ref, row_num):
    errors = []
    
    for col, name, ftype, req, _ in STORE_FIELDS:
        val = row[col] if col < len(row) else None
        ok, msg = True, None
        
        if ftype == 'enum':
            eset = STORE_FIELDS[[f[0] for f in STORE_FIELDS].index(col)][4]
            ok, msg = chk_enum(name, val, eset)
        elif ftype == 'uint':
            ok, msg = chk_uint(name, val, req)
        elif ftype == 'phone':
            ok, msg = chk_phone(name, val, req)
        elif ftype == 'code':
            plan = row[4] if 4 < len(row) else ''
            ok, msg = chk_code(name, val, plan, req)
        elif ftype == 'ref':
            label = STORE_FIELDS[[f[0] for f in STORE_FIELDS].index(col)][4]
            rset = ref.geo_codes if label == 'geo' else ref.biz_codes
            ok, msg = chk_ref(name, val, rset, req, label)
        elif ftype == 'chan':
            ok, msg = chk_chan(name, val, req)
        elif ftype == 'pos':
            ok, msg = chk_pos(name, val)
        elif ftype == 'text':
            if req and is_blank(val):
                ok, msg = False, '该字段必填'
        
        if not ok:
            errors.append({
                'row': row_num, 'col': col, 'name': name,
                'value': safeStr(val), 'desc': msg, 'sheet': '通路'
            })
    
    # 勾稽关系: K = 价位段 = 品牌
    p_sum = sum(to_int(row[i]) or 0 for i in (19, 20, 21) if i < len(row))
    b_sum = sum(to_int(row[i]) or 0 for i in (22, 23, 24, 25, 26, 27) if i < len(row))
    k_val = to_int(row[10]) if 10 < len(row) else None
    is_closed = len(row) > 12 and row[12] == '闭店'
    
    if not (is_closed and p_sum == 0 and b_sum == 0):
        if p_sum != b_sum:
            errors.append({
                'row': row_num, 'col': 19, 'name': '价位段≠品牌',
                'value': f'{p_sum}≠{b_sum}',
                'desc': f'价位段({p_sum})≠品牌合计({b_sum})', 'sheet': '通路'
            })
        if k_val is not None and k_val != p_sum:
            errors.append({
                'row': row_num, 'col': 10, 'name': 'capa≠价位段',
                'value': f'{k_val}≠{p_sum}',
                'desc': f'门店capa({k_val})≠价位段之和({p_sum})', 'sheet': '通路'
            })
    
    return errors


def validate_customer_row(row, row_num):
    errors = []
    
    for col, name, ftype, req, _ in CUSTOMER_FIELDS:
        val = row[col] if col < len(row) else None
        ok, msg = True, None
        
        if ftype == 'enum':
            eset = CUSTOMER_FIELDS[[f[0] for f in CUSTOMER_FIELDS].index(col)][4]
            ok, msg = chk_enum(name, val, eset)
        elif ftype == 'uint':
            ok, msg = chk_uint(name, val, req)
        elif ftype == 'phone':
            ok, msg = chk_phone(name, val, req)
        elif ftype == 'text':
            if req and is_blank(val):
                ok, msg = False, '该字段必填'
        
        if not ok:
            errors.append({
                'row': row_num, 'col': col, 'name': name,
                'value': safeStr(val), 'desc': msg, 'sheet': '客户'
            })
    
    return errors


# ==================== 文件校验 ====================

def read_reference_data(wb):
    ref = RefData()
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = [cell.value for cell in ws[1]]
        is_geo = any('地理编码' in str(h) or '地理信息编码' in str(h) for h in headers if h)
        is_biz = any('商圈编码' in str(h) for h in headers if h)
        if not is_geo and not is_biz:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if cell:
                    s = safeStr(cell)
                    if s:
                        if is_geo:
                            ref.geo_codes.add(s)
                        if is_biz:
                            ref.biz_codes.add(s)
    return ref


def validate_file(file_path, progress_callback=None):
    """校验单个文件，返回 (store_errors, cust_errors, filename)"""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ref = read_reference_data(wb)
    
    store_errors = []
    cust_errors = []
    
    # 校验通路门店
    if '通讯通路门店' in wb.sheetnames:
        ws = wb['通讯通路门店']
        rows = list(ws.iter_rows(values_only=True))
        for i in range(3, len(rows)):
            store_errors.extend(validate_store_row(rows[i], ref, i + 1))
            if progress_callback and i % 100 == 0:
                progress_callback('store', i, len(rows))
    
    # 校验通讯客户沙盘
    if '通讯客户沙盘' in wb.sheetnames:
        ws = wb['通讯客户沙盘']
        rows = list(ws.iter_rows(values_only=True))
        for i in range(3, len(rows)):
            cust_errors.extend(validate_customer_row(rows[i], i + 1))
            if progress_callback and i % 100 == 0:
                progress_callback('customer', i, len(rows))
    
    wb.close()
    return store_errors, cust_errors


def write_results_to_excel(file_path, store_errors, cust_errors, progress_callback=None):
    """将校验结果写入原文件最后列"""
    wb = openpyxl.load_workbook(file_path)
    
    # 样式
    YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    BOLD_FONT = Font(bold=True)
    
    for sheet_name, errors in [('通讯通路门店', store_errors), ('通讯客户沙盘', cust_errors)]:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        max_col = ws.max_column
        check_col = get_column_letter(max_col + 1)
        
        # 表头
        ws.cell(row=2, column=max_col + 1, value='校验结果').font = BOLD_FONT
        ws.cell(row=2, column=max_col + 1).fill = YELLOW_FILL
        
        # 按行聚合
        row_map = defaultdict(list)
        for e in errors:
            row_map[e['row']].append(e)
        
        # 设置列宽
        ws.column_dimensions[check_col].width = 80
        
        for row_num, errs in row_map.items():
            cell = ws.cell(row=row_num, column=max_col + 1)
            text = ' | '.join(f"行{e['row']} [{e['name']}] '{e['value'][:20]}' → {e['desc']}" for e in errs)
            if len(text) > 490:
                text = text[:490]
            cell.value = f'❌ {text}'
            cell.fill = RED_FILL
        
        # 正确行
        for row_idx in range(4, ws.max_row + 1):
            if row_idx not in row_map:
                cell = ws.cell(row_idx, column=max_col + 1)
                cell.value = '✅ 正确'
                cell.fill = GREEN_FILL
    
    output = file_path.replace('.xlsx', '_校验结果.xlsx')
    wb.save(output)
    wb.close()
    return output


# ==================== GUI ====================

class SandboxValidatorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title('通讯沙盘校验工具 v5.0')
        self.root.geometry('950x650')
        self.root.resizable(True, True)
        
        self.files = []
        self.results = {}
        self.running = False
        
        self._build_ui()
    
    def _build_ui(self):
        # 顶部按钮栏
        top_frame = ttk.Frame(self.root, padding=10)
        top_frame.pack(fill=tk.X)
        
        ttk.Button(top_frame, text='📂 选择文件', command=self.select_files, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text='📁 选择文件夹', command=self.select_folder, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text='🗑️ 清空', command=self.clear_files, width=10).pack(side=tk.LEFT, padx=5)
        
        self.start_btn = ttk.Button(top_frame, text='▶ 开始校验', command=self.start_validation, width=15)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(top_frame, text='💾 导出报告', command=self.export_report, width=12).pack(side=tk.RIGHT, padx=5)
        
        # 文件列表
        list_frame = ttk.LabelFrame(self.root, text='待校验文件', padding=5)
        list_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.file_listbox = tk.Listbox(list_frame, height=6, font=('Consolas', 10))
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.file_listbox.yview)
        self.file_listbox.configure(yscrollcommand=scrollbar.set)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 进度条
        progress_frame = ttk.Frame(self.root, padding=(10, 5))
        progress_frame.pack(fill=tk.X)
        
        self.progress_label = ttk.Label(progress_frame, text='就绪')
        self.progress_label.pack(side=tk.LEFT)
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate', length=400)
        self.progress_bar.pack(side=tk.RIGHT, padx=10)
        
        # 结果树
        result_frame = ttk.LabelFrame(self.root, text='校验结果', padding=5)
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        columns = ('文件', 'Sheet', '行', '字段', '值', '描述')
        self.tree = ttk.Treeview(result_frame, columns=columns, show='headings', height=12)
        
        for col in columns:
            self.tree.heading(col, text=col)
        
        self.tree.column('文件', width=150, minwidth=100)
        self.tree.column('Sheet', width=80)
        self.tree.column('行', width=50)
        self.tree.column('字段', width=100)
        self.tree.column('值', width=80)
        self.tree.column('描述', width=250)
        
        tree_scroll = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 双击打开文件
        self.tree.bind('<Double-1>', self._on_double_click)
        
        # 状态栏
        self.status_bar = ttk.Label(self.root, text='  就绪 | 文件: 0 | 错误: 0', relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM)
    
    def select_files(self):
        files = filedialog.askopenfilenames(
            title='选择 Excel 文件',
            filetypes=[('Excel 文件', '*.xlsx'), ('所有文件', '*.*')]
        )
        for f in files:
            if f not in self.files:
                self.files.append(f)
                self.file_listbox.insert(tk.END, f)
        self._update_status()
    
    def select_folder(self):
        folder = filedialog.askdirectory(title='选择文件夹')
        if folder:
            for f in Path(folder).rglob('*.xlsx'):
                fp = str(f)
                if fp not in self.files and '_校验结果' not in fp:
                    self.files.append(fp)
                    self.file_listbox.insert(tk.END, fp)
            self._update_status()
    
    def clear_files(self):
        self.files.clear()
        self.file_listbox.delete(0, tk.END)
        self._update_status()
    
    def _update_status(self):
        self.status_bar.config(text=f'  就绪 | 文件: {len(self.files)} | 错误: {len(self.results)}')
    
    def start_validation(self):
        if not self.files:
            messagebox.showwarning('提示', '请先选择文件或文件夹')
            return
        
        if self.running:
            return
        
        self.running = True
        self.start_btn.config(state=tk.DISABLED)
        self.tree.delete(*self.tree.get_children())
        self.results.clear()
        
        thread = threading.Thread(target=self._run_validation, daemon=True)
        thread.start()
    
    def _run_validation(self):
        total = len(self.files)
        all_errors = []
        
        for idx, fp in enumerate(self.files):
            self._set_progress(f'正在校验: {os.path.basename(fp)} ({idx+1}/{total})', 
                                (idx / total) * 100)
            
            try:
                store_errs, cust_errs = validate_file(fp, self._file_progress)
                
                for e in store_errs:
                    e['file'] = fp
                for e in cust_errs:
                    e['file'] = fp
                
                errors = store_errs + cust_errs
                all_errors.extend(errors)
                self.results[fp] = {
                    'store': len(store_errs),
                    'customer': len(cust_errs),
                    'total': len(store_errs) + len(cust_errs)
                }
                
                # 写入结果文件
                out_path = write_results_to_excel(fp, store_errs, cust_errs)
                
                # 在结果树中显示
                for e in errors:
                    self._insert_result(os.path.basename(fp), '通讯通路门店' if e['sheet']=='通路' else '通讯客户沙盘',
                                       e['row'], e['name'], e['value'], e['desc'])
                
                self._insert_summary(fp, len(store_errs), len(cust_errs), out_path)
                
            except Exception as e:
                self._insert_result(os.path.basename(fp), 'ERROR', '-', '-', '-', str(e))
            
            self.root.update_idletasks()
        
        self._set_progress(f'完成: {total} 个文件, 共 {len(all_errors)} 条错误', 100)
        self.status_bar.config(text=f'  完成 | 文件: {total} | 错误: {len(all_errors)}')
        self.running = False
        self.start_btn.config(state=tk.NORMAL)
        
        if total == 1 and all_errors == 0:
            messagebox.showinfo('完成', '校验通过！所有数据正确。')
    
    def _file_progress(self, sheet_type, current, total):
        self.root.after(0, lambda: self._set_progress(
            f'{sheet_type}: {current}/{total}', None))
    
    def _set_progress(self, label, value):
        self.root.after(0, lambda: self.progress_label.config(text=label))
        if value is not None:
            self.root.after(0, lambda: self.progress_bar.configure(value=value))
    
    def _insert_result(self, filename, sheet, row, field, value, desc):
        self.root.after(0, lambda: self.tree.insert(
            '', tk.END, values=(filename, sheet, row, field, str(value)[:30], desc)
        ))
    
    def _insert_summary(self, fp, store_errs, cust_errs, out_path):
        summary = f'[完成] {os.path.basename(fp)} - 通路 {store_errs} 条 + 客户 {cust_errs} 条 = {store_errs + cust_errs} 条错误'
        self.root.after(0, lambda: self.tree.insert(
            '', tk.END, values=(os.path.basename(fp), '完成', '-', '-', f'{store_errs+cust_errs} 条错误', out_path)
        ))
    
    def _on_double_click(self, event):
        sel = self.tree.selection()
        if sel:
            item = self.tree.item(sel[0])
            filepath = item['values'][0]
            if os.path.exists(filepath):
                os.startfile(os.path.dirname(os.path.abspath(filepath)))
    
    def export_report(self):
        if not self.results:
            messagebox.showwarning('提示', '没有校验结果可导出')
            return
        
        fp = filedialog.asksaveasfilename(
            title='保存报告',
            defaultextension='.txt',
            filetypes=[('文本文件', '*.txt'), ('JSON', '*.json'), ('所有文件', '*.*')]
        )
        if not fp:
            return
        
        if fp.endswith('.json'):
            with open(fp, 'w', encoding='utf-8') as f:
                json.dump(self.results, f, ensure_ascii=False, indent=2, default=str)
        else:
            with open(fp, 'w', encoding='utf-8') as f:
                f.write(f'通讯沙盘校验报告 - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
                f.write('=' * 60 + '\n\n')
                for filepath, data in self.results.items():
                    f.write(f'{os.path.basename(filepath)}\n')
                    f.write(f'  通路: {data["store"]} 条错误\n')
                    f.write(f'  客户: {data["customer"]} 条错误\n')
                    f.write(f'  总计: {data["total"]} 条错误\n\n')
        
        messagebox.showinfo('完成', f'报告已保存到:\n{fp}')


class ScanDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title('扫描选项')
        self.geometry('300x150')
        self.resizable(False, False)
        
        ttk.Label(self, text='扫描进度', padding=10).pack()
        
        self.progress = ttk.Progressbar(self, mode='indeterminate', length=250)
        self.progress.pack(padx=10, pady=5)
        self.progress.start(10)
        
        self.label = ttk.Label(self, text='正在扫描...')
        self.label.pack(pady=5)


def main():
    root = tk.Tk()
    app = SandboxValidatorGUI(root)
    root.mainloop()


if __name__ == '__main__':
    main()
