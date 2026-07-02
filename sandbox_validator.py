# -*- coding: utf-8 -*-
"""
通讯沙盘填写规范校验工具 v3.1
单文件，Mac/Win 离线

依赖: pip install openpyxl
"""

import os
import re
import sys
import time
import threading
from collections import defaultdict
from tkinter import (Tk, Frame, Label, Button, Text, Scrollbar,
                     filedialog, messagebox, StringVar)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill


# ============================================================
# 基础
# ============================================================

def to_int(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v or abs(v) == float('inf')):
            return None
        return int(v)
    s = str(v).strip().replace(",", "").replace("，", "")
    if not s or s == "nan":
        return None
    try:
        return int(float(s))
    except:
        return None

def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")

def safe_str(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s == "nan" else s

def clean_phone(v):
    return "" if v is None else re.sub(r'[^0-9]', '', str(v))

RED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
BOLD_RED = Font(bold=True, color="CC0000")

class Err:
    """校验错误"""
    __slots__ = ['sheet','row','col','name','value','desc','etype']
    def __init__(self, sheet, row, col, name, value, desc, etype="错误"):
        self.sheet = sheet
        self.row = row
        self.col = col
        self.name = name
        self.value = value
        self.desc = desc
        self.etype = etype

    def __str__(self):
        return (f"行{self.row} [{self.name}] "
                f"'{safe_str(self.value)[:20] or '(空)'}' → {self.desc}")


# ============================================================
# 字段定义：通讯通路门店 (67列 0-66, 校验列=65)
# ============================================================

STORE_FIELDS = [
    # --- 位置 (0-3) ---
    (0,  "省份",       "enum",  True,  {"海南省","广东省","广西壮族自治区","云南省","四川省","河北省","山西省","辽宁省","吉林省","黑龙江省","江苏省","浙江省","安徽省","福建省","山东省","黑龙江省","新疆维吾尔自治区","宁夏回族自治区","内蒙古自治区","广西壮族自治区","陕西省","甘肃省","青海省","贵州省","山东省","北京市","上海市","天津市","重庆市","台湾省","香港特别行政区","澳门特别行政区"}),
    (1,  "城市",       "text",  True,  None),
    (2,  "区县",       "text",  True,  None),
    (3,  "乡镇",       "text",  True,  None),
    # --- 门店 (4-18) ---
    (4,  "进度",       "enum",  True,  {"已存在","无规划","洽谈中"}),
    (5,  "编码",       "scode", True,  None),
    (6,  "名称",       "text",  True,  None),
    (7,  "地理编码",    "ref",   True,  "geo"),
    (8,  "商圈编码",    "ref",   True,  "biz"),
    (9,  "类型",       "enum",  True,  {"NKA综合卖场","体验店","运营商厅店","专卖店","其他","华为授权体验店","小米体验店","综合卖场","耀生活","苹果体验店","荣耀体验店","荣耀授权专卖店","其他品牌专卖店"}),
    (10, "capa",       "uint",  True,  None),
    (11, "面积",       "uint_", False,  None),
    (12, "状态",       "enum",  True,  {"不合作","营业中","闭店","装修中"}),
    (13, "运营商",     "enum",  True,  {"公开","公开&三网","排它-移动","排它-电信","排它-联通","非排它-移动电信","非排它-移动联通","非排它-电信联通","排它-移动"}),
    (14, "联系人",     "text",  True,  None),
    (15, "联系电话",   "phone", True,  None),
    (16, "地址",       "text",  True,  None),
    (17, "渠道编码",   "chan",  True,  None),
    (18, "客户名称",   "text",  True,  None),
    # --- 价位段 (19-21) ---
    (19, "0-2500",     "uint",  True,  None),
    (20, "2500-4000",  "uint",  True,  None),
    (21, ">4000",      "uint",  True,  None),
    # --- 品牌手机 (22-27) ---
    (22, "荣耀",       "uint",  True,  None),
    (23, "华为",       "uint",  True,  None),
    (24, "小米",       "uint",  True,  None),
    (25, "VIVO",      "uint",  True,  None),
    (26, "苹果",       "uint",  True,  None),
    (27, "OPPO",      "uint",  True,  None),
    # --- 全场景 (28-34) ---
    (28, "荣耀PC",    "uint",  True,  None),
    (29, "华为PC",    "uint",  True,  None),
    (30, "荣耀平板",  "uint",  True,  None),
    (31, "华为平板",  "uint",  True,  None),
    (32, "PC",        "uint",  True,  None),
    (33, "平板",      "uint",  True,  None),
    (34, "手表",      "uint",  True,  None),
    # --- 品牌阵地 x6 (35-52) ---
    (35, "荣耀阵地数",  "uint_", False, None),
    (36, "荣耀阵地位",  "pos",  False, None),
    (37, "荣耀门头数",  "uint_", False, None),
    (38, "华为阵地数",  "uint_", False, None),
    (39, "华为阵地位",  "pos",  False, None),
    (40, "华为门头数",  "uint_", False, None),
    (41, "小米阵地数",  "uint_", False, None),
    (42, "小米阵地位",  "pos",  False, None),
    (43, "小米门头数",  "uint_", False, None),
    (44, "VIVO阵地数", "uint_", False, None),
    (45, "VIVO阵地位", "pos",  False, None),
    (46, "VIVO门头数", "uint_", False, None),
    (47, "苹果营地数",  "uint_", False, None),
    (48, "苹果阵地位",  "pos",  False, None),
    (49, "苹果门头数",  "uint_", False, None),
    (50, "OPPO阵地数", "uint_", False, None),
    (51, "OPPO阵地位", "pos",  False, None),
    (52, "OPPO门头数", "uint_", False, None),
    # --- 促销员 (53-58) ---
    (53, "荣耀促销员",  "uint",  True,  None),
    (54, "华为促销员",  "uint",  True,  None),
    (55, "小米促销员",  "uint",  True,  None),
    (56, "VIVO促销员", "uint",  True,  None),
    (57, "苹果促销员",  "uint",  True,  None),
    (58, "OPPO促销员", "uint",  True,  None),
    # --- 其他 (59-64) ---
    (59, "客户员工",    "uint",  True,  None),
    (60, "店端人数",    "uint",  True,  None),
    (61, "国补",        "enum",  True,  {"是","否"}),
    (62, "城乡类别",    "enum",  True,  {"主城区","非主城区"}),
    (63, "街边/MALL",  "enum",  True,  {"街边店","MALL店"}),
    (64, "备注",        "text",  False, None),
]

STORE_CHECK_COL = 65  # 0-indexed 校验结果写入列

# ============================================================
# 字段定义：通讯客户沙盘 (58列 0-57, 校验列=59)
# ============================================================

# 通讯客户沙盘：仅校验 E(4)、F(5)、I-T(8-19)、AQ(42) 列
CUSTOMER_FIELDS = [
    (4,  "跨省份",     "enum",  True,  {"是","否"}),
    (5,  "跨地市",     "enum",  True,  {"是","否"}),
    (8,  "圈层",       "text",  True,  None),
    (9,  "华为属性",   "enum",  True,  {"金种子","TOP368","TOP1000","TOP1500","NDKA+","TOP2000","TOP5000","其他"}),
    (10, "OPPO层级",   "enum",  True,  {"T+","T","S","A","其他"}),
    (11, "vivo层级",   "enum",  True,  {"黑金+","黑金","钻石","铂金","其他"}),
    (12, "小米层级",   "enum",  True,  {"蓝血Ultra","蓝血Plus","蓝血","金牌","卓越Ultra","卓越Plus","卓越","优秀","其他"}),
    (13, "运营商属性", "enum",  True,  {"公开","公开&三网","排它-移动","排它-电信","排它-联通","非排它-移动电信","非排它-移动联通","非排它-电信联通"}),
    (14, "友商代理",   "enum",  True,  {"是","否"}),
    (15, "车授权",     "enum",  True,  {"是","否"}),
    (16, "老板名",     "text",  True,  None),
    (17, "老板电话",   "phone", True,  None),
    (18, "操盘手名",   "text",  True,  None),
    (19, "操盘手电话",  "phone", True,  None),
    (42, "0-2500容量", "uint",  True,  None),
]

CUSTOMER_CHECK_COL = 43


# ============================================================
# 校验规则
# ============================================================

def chk_enum(name, val, eset):
    if is_blank(val):
        return False, "该字段必填"
    sv = safe_str(val)
    if sv not in eset:
        return False, f"无效值，允许: {','.join(sorted(eset))}"
    return True, None

def chk_uint(name, val, req=True):
    if is_blank(val):
        return (False, "该字段必填") if req else (True, None)
    iv = to_int(val)
    if iv is None:
        return False, "非有效整数"
    if iv < 0 or iv > 999999:
        return False, f"超出范围(0-999999): {iv}"
    return True, None

def chk_phone(name, val, req=True):
    if is_blank(val):
        return (False, "该字段必填") if req else (True, None)
    s = clean_phone(val)
    if not re.match(r'^1[3-9]\d{9}$', s):
        return False, "无效手机号"
    return True, None

def chk_code(val, plan, req=True):
    if is_blank(val):
        if req and plan == "已存在":
            return False, "已存在门店必须填写编码"
        return (False, "该字段必填") if req else (True, None)
    s = safe_str(val)
    if s in ("无规划", "洽谈中", "/", "无"):
        return True, None
    if re.match(r'^[A-Za-z]', s) or re.match(r'^46\d{6,}$', s) or (s.isdigit() and len(s) >= 8):
        return True, None
    return False, "编码格式异常"

def chk_ref(val, ref_set, req=True, label=""):
    if is_blank(val):
        return (False, "该字段必填") if req else (True, None)
    if safe_str(val) not in ref_set:
        return False, f"{label}编码不存在"
    return True, None

def chk_chan(val, req=True):
    if is_blank(val):
        return (False, "该字段必填") if req else (True, None)
    s = safe_str(val).replace("®", "").replace("(R)", "").strip()
    if s == "无" or s.isdigit():
        return True, None
    return False, "渠道编码应为纯数字或'无'"

def chk_pos(val):
    if is_blank(val):
        return True, None
    sv = safe_str(val)
    if sv in ("", "0", "无", "无阵地"):
        return True, None
    valid = {"TOP1","TOP2","TOP3","其他","无地面","无地"}
    if sv not in valid:
        return False, "无效枚举"
    return True, None


# ============================================================
# 核心校验函数
# ============================================================

def validate_store_row(row, ref):
    """校验通路门店一行数据"""
    errors = []

    # 跳过完全空行（所有值都为 None）
    vals = [v for v in row.values() if v is not None and str(v).strip() != ""]
    if not vals:
        return errors

    for col, name, ftype, req, extra in STORE_FIELDS:
        val = row.get(col)

        if ftype == "enum":
            ok, msg = chk_enum(name, val, extra)
        elif ftype in ("uint",):
            ok, msg = chk_uint(name, val, req)
        elif ftype == "phone":
            ok, msg = chk_phone(name, val, req)
        elif ftype == "scode":
            plan = row.get(4)
            ok, msg = chk_code(val, plan, req)
        elif ftype == "ref":
            rkey = extra  # "geo" or "biz"
            ref_set = ref.get(rkey + "_codes", set())
            ok, msg = chk_ref(val, ref_set, req, label=rkey)
        elif ftype == "chan":
            ok, msg = chk_chan(val, req)
        elif ftype == "pos":
            ok, msg = chk_pos(val)
        elif ftype == "text":
            ok, msg = (True, None) if not req or not is_blank(val) else (False, "该字段必填")
        else:
            ok, msg = True, None

        if not ok:
            errors.append(Err("通讯通路门店", row.get("_r"), col, name, val, msg))

    # 勾稽: 价位段求和 = 品牌求和
    p = sum([to_int(row.get(i)) or 0 for i in (19, 20, 21)])
    b = sum([to_int(row.get(i)) or 0 for i in (22, 23, 24, 25, 26, 27)])
    is_closed = safe_str(row.get(12)) == "闭店"
    if p != b:
        if not (is_closed and p == 0 and b == 0):
            errors.append(Err("通讯通路门店", row.get("_r"), 19,
                "价位段≠品牌", f"{p}≠{b}", f"价位段({p})≠品牌合计({b})", "逻辑"))

    # 勾稽：TOP位唯一（TOP1/TOP2/TOP3 每行只能出现一次）
    top_cols = [36, 39, 42, 45, 48, 51]  # AK/AN/AQ/AT/AW/BA(位置列)
    top_vals = []
    for c in top_cols:
        v = safe_str(row.get(c) or "")
        if v in ("TOP1", "TOP2", "TOP3"):
            top_vals.append(v)
    if len(top_vals) != len(set(top_vals)):
        errors.append(Err("通讯通路门店", row.get("_r"), 36,
            "TOP位重复", f"{top_vals}", "TOP1/TOP2/TOP3 不可重复", "逻辑"))

    return errors


def validate_customer_row(row, ftype_overrides=None):
    """校验通讯客户沙盘一行数据，ftype_overrides 可跳过指定列的基础校验"""
    errors = []
    if ftype_overrides is None:
        ftype_overrides = {}

    for col, name, ftype, req, extra in CUSTOMER_FIELDS:
        # 如果该列被标记为跳过基础校验
        if col in ftype_overrides:
            continue
        val = row.get(col)

        if ftype == "enum":
            ok, msg = chk_enum(name, val, extra)
        elif ftype == "uint":
            ok, msg = chk_uint(name, val, req)
        elif ftype == "phone":
            ok, msg = chk_phone(name, val, req)
        elif ftype == "chan":
            ok, msg = chk_chan(val, req)
        elif ftype == "text":
            ok, msg = (True, None) if not req or not is_blank(val) else (False, "该字段必填")
        else:
            ok, msg = True, None

        if not ok:
            errors.append(Err("通讯客户沙盘", row.get("_r"), col, name, val, msg))

    return errors


# ============================================================
# 引用数据
# ============================================================

def read_reference(wb_ro):
    """从只读工作簿读取引用数据"""
    ref = {"geo_codes": set(), "biz_codes": set()}
    for sn in wb_ro.sheetnames:
        ws = wb_ro[sn]
        if "地理" in sn or "GIS" in sn.upper():
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if cell:
                        ref["geo_codes"].add(safe_str(cell))
        elif "商圈" in sn:
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if cell:
                        ref["biz_codes"].add(safe_str(cell))
    return ref


# ============================================================
# 批量校验（使用 iter_rows 加速）
# ============================================================

def validate_store_all(wb_ro, ref):
    """校验通讯通路门店"""
    ws = wb_ro["通讯通路门店"]
    col_indices = [f[0] for f in STORE_FIELDS]
    max_col = max(col_indices) + 1
    all_errs = []
    for row_idx, raw in enumerate(ws.iter_rows(min_row=4, max_col=max_col, values_only=True), start=4):
        row_data = {"_r": row_idx}
        for i, col in enumerate(col_indices):
            row_data[col] = raw[i] if i < len(raw) else None
        all_errs.extend(validate_store_row(row_data, ref))
    return all_errs

def validate_customer_all(wb_ro):
    """校验通讯客户沙盘"""
    ws = wb_ro["通讯客户沙盘"]
    col_indices = [f[0] for f in CUSTOMER_FIELDS]
    max_col = max(col_indices) + 1
    all_errs = []
    for row_idx, raw in enumerate(ws.iter_rows(min_row=4, max_col=max_col, values_only=True), start=4):
        row_data = {"_r": row_idx}
        for i, col in enumerate(col_indices):
            row_data[col] = raw[i] if i < len(raw) else None
        all_errs.extend(validate_customer_row(row_data))
    return all_errs
    return all_errs


# ============================================================
# 写入结果
# ============================================================

def write_errors(ws, errs, _):
    """将校验结果写入数据最后一列之后"""
    from openpyxl.utils import get_column_letter

    # 动态确定输出列（数据最后+1列）
    data_max_col = 0
    for col_idx in range(ws.max_column, 0, -1):
        val = ws.cell(row=2, column=col_idx).value
        if val is not None:
            data_max_col = col_idx
            break

    start_col = data_max_col + 1 if data_max_col > 0 else 2

    merged_ranges = set()
    if hasattr(ws, 'merged_cells'):
        for mr in ws.merged_cells.ranges:
            merged_ranges.add((mr.min_row, mr.min_col, mr.max_row, mr.max_col))

    # 清除旧结果
    for row in range(1, ws.max_row + 1):
        for col in range(start_col, ws.max_column + 1):
            is_m = any(r1 <= row <= r2 and c1 <= col <= c2 for r1, c1, r2, c2 in merged_ranges)
            if not is_m:
                c = ws.cell(row=row, column=col)
                c.value = None
                c.fill = PatternFill()

    # 表头
    ws.cell(row=2, column=start_col).value = "校验结果"
    ws.cell(row=2, column=start_col).font = Font(bold=True)
    ws.column_dimensions[get_column_letter(start_col)].width = 80

    # 写入
    by_row = defaultdict(list)
    for e in errs:
        by_row[e.row].append(e)

    for ri, errs_list in by_row.items():
        c = ws.cell(row=ri, column=start_col)
        text = " | ".join(str(e) for e in errs_list)
        c.value = f"❌ {text[:490]}"
        c.font = BOLD_RED
        c.fill = RED_FILL

    return len(by_row)


# ============================================================
# GUI 应用
# ============================================================

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("通讯沙盘填写规范校验工具 v3.1")
        self.root.geometry("1000x720")
        self.file_paths = []
        self._build()

    def _build(self):
        top = Frame(self.root, padx=12, pady=10)
        top.pack(fill="x")
        Label(top, text="通讯沙盘填写规范校验工具",
              font=("Microsoft YaHei", 16, "bold")).pack(side="left")

        bf = Frame(self.root, padx=12, pady=5)
        bf.pack(fill="x")
        Button(bf, text="📂 选文件", command=self.sel_files,
               bg="#4CAF50", fg="white",
               font=("Microsoft YaHei", 11), padx=15, pady=5).pack(side="left", padx=5)
        Button(bf, text="📁 选文件夹", command=self.sel_folder,
               bg="#2196F3", fg="white",
               font=("Microsoft YaHei", 11), padx=15, pady=5).pack(side="left", padx=5)
        self.file_label = Label(bf, text="未选择文件",
                               fg="gray", font=("Microsoft YaHei", 10))
        self.file_label.pack(side="left", padx=10)

        lf = Frame(self.root, padx=12, pady=5)
        lf.pack(fill="both", expand=True)
        self.txt = Text(lf, font=("Consolas", 10), bg="#1e1e1e",
                       fg="#00ff00", insertbackground="white", wrap="word")
        sb = Scrollbar(lf, command=self.txt.yview)
        self.txt.config(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.txt.pack(fill="both", expand=True)

        bot = Frame(self.root, padx=12, pady=10)
        bot.pack(fill="x")
        Button(bot, text="▶ 开始校验", command=self.run,
               bg="#FF5722", fg="white",
               font=("Microsoft YaHei", 12, "bold"),
               padx=20, pady=8).pack(side="left", padx=5)
        Button(bot, text="🧹 清空日志", command=self.clear,
               bg="#607D8B", fg="white",
               font=("Microsoft YaHei", 11),
               padx=15, pady=5).pack(side="left", padx=5)
        Button(bot, text="📋 导出报告", command=self.export,
               bg="#9C27B0", fg="white",
               font=("Microsoft YaHei", 11),
               padx=15, pady=5).pack(side="left", padx=5)

        self.status = StringVar(value="就绪 | Mac/Win 离线 | pip install openpyxl")
        Label(self.root, textvariable=self.status, bd=1, relief="sunken",
              anchor="w", font=("Microsoft YaHei", 9)).pack(side="bottom", fill="x")

    def log(self, msg):
        self.txt.insert("end", msg + "\n")
        self.txt.see("end")
        self.root.update()

    def clear(self):
        self.txt.delete("1.0", "end")

    def sel_files(self):
        paths = filedialog.askopenfilenames(
            title="选择 Excel", filetypes=[("Excel", "*.xlsx"), ("所有", "*.*")])
        if paths:
            self.file_paths = list(paths)
            self.file_label.config(
                text=f"已选 {len(self.file_paths)} 个文件", fg="black")

    def sel_folder(self):
        folder = filedialog.askdirectory(title="选择文件夹")
        if folder:
            self.file_paths = []
            for dp, _, fn_list in os.walk(folder):
                for f in fn_list:
                    if f.endswith(".xlsx") and not f.startswith("~"):
                        self.file_paths.append(os.path.join(dp, f))
            self.file_label.config(
                text=f"已扫描 {len(self.file_paths)} 个文件", fg="black")

    def run(self):
        if not self.file_paths:
            messagebox.showwarning("提示", "请先选择文件")
            return
        threading.Thread(target=self._process, daemon=True).start()

    def _process(self):
        t00 = time.time()
        total_errs = total_ok = 0

        for i, fp in enumerate(self.file_paths):
            self.log(f"\n{'='*60}")
            self.log(f"[{i+1}/{len(self.file_paths)}] {os.path.basename(fp)}")
            self.log(f"{'='*60}")
            t0 = time.time()

            try:
                # Step 1: 只读模式加载（快速）
                wb_ro = openpyxl.load_workbook(fp, read_only=True, data_only=True)
                ref = read_reference(wb_ro)
                self.log(f"引用: {len(ref['geo_codes'])} 地理, "
                         f"{len(ref['biz_codes'])} 商圈")

                # Step 2: 校验
                store_errs = []
                if "通讯通路门店" in wb_ro.sheetnames:
                    store_errs = validate_store_all(wb_ro, ref)
                    self.log(f"通路门店: {len(store_errs)} 错误")

                cust_errs = []
                if "通讯客户沙盘" in wb_ro.sheetnames:
                    cust_errs = validate_customer_all(wb_ro)
                    self.log(f"客户沙盘: {len(cust_errs)} 错误")

                wb_ro.close()

                # Step 3: 写模式加载（仅写校验列）
                wb = openpyxl.load_workbook(fp, data_only=False)

                if store_errs:
                    write_errors(wb["通讯通路门店"], store_errs, STORE_CHECK_COL)
                if cust_errs:
                    write_errors(wb["通讯客户沙盘"], cust_errs, CUSTOMER_CHECK_COL)

                file_errs = len(store_errs) + len(cust_errs)
                if file_errs == 0:
                    total_ok += 1
                    self.log("✅ 无错误")
                else:
                    total_errs += file_errs

                # Step 4: 保存
                try:
                    wb.save(fp)
                    self.log(f"已保存 ({time.time()-t0:.1f}s)")
                except:
                    alt = fp.replace(".xlsx", "_校验结果.xlsx")
                    try:
                        wb.save(alt)
                        self.log(f"已另存: {os.path.basename(alt)}")
                    except Exception as e2:
                        self.log(f"保存失败: {e2}")

                wb.close()

            except Exception as e:
                self.log(f"❌ 处理失败: {e}")

        elapsed = time.time() - t00
        self.log(f"\n{'='*60}")
        self.log(f"完成: {len(self.file_paths)} 文件, "
                 f"{total_ok} 文件无问题, 共 {total_errs} 条错误")
        self.log(f"耗时: {elapsed:.1f}s")
        self.status.set(f"完成: {total_errs} 条错误, {elapsed:.1f}s")

    def export(self):
        content = self.txt.get("1.0", "end").strip()
        if not content:
            messagebox.showinfo("提示", "无内容")
            return
        path = filedialog.asksaveasfilename(
            title="导出报告", defaultextension=".txt",
            filetypes=[("文本", "*.txt")])
        if path:
            with open(path, "w", encoding="utf-8") as f:
                f.write(content)
            messagebox.showinfo("完成", f"已保存: {path}")


def main():
    root = Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
